import logging
from datetime import timedelta
from urllib.parse import urlencode

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import SetPasswordForm
from django.contrib.auth.views import LoginView, PasswordChangeView
from django.core.cache import cache
from django.core.exceptions import PermissionDenied
from django.core.mail import send_mail
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Count, Q
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from info.decorators import (
    assert_teaches,
    owns_assign,
    owns_attendance_class,
    owns_marks_class,
    owns_teacher_id,
    teacher_required,
)
from info.forms import (
    BulkFeeForm,
    ErpLoginForm,
    ExtraClassForm,
    FeeForm,
    FeeTransactionForm,
    MarksEntryForm,
    NoticeForm,
    PasswordResetRequestForm,
    PasswordResetVerifyForm,
    ProfileForm,
    StudentForm,
    SupportRequestForm,
    TeacherForm,
)
from info.imports import (
    MAX_ROWS,
    STUDENT_COLUMNS,
    TEACHER_COLUMNS,
    ImportError_,
    commit,
    read_rows,
    validate,
)
from info.reports import payment_receipt, report_card
from info.services import (
    SessionNotMarkable,
    attendance_rows,
    submit_attendance,
    submit_marks,
)

from .models import (
    CIE_MAX,
    CLASS_CANCELLED,
    CLASS_PENDING,
    DAYS_OF_WEEK,
    SEE_ELIGIBILITY_CIE,
    SEE_MAX,
    Assign,
    AssignTime,
    Attendance,
    AttendanceClass,
    AttendanceTotal,
    AuditLog,
    Class,
    Course,
    Dept,
    Fee,
    FeeTransaction,
    Marks,
    MarksClass,
    Notice,
    NoticeRead,
    PasswordResetOTP,
    Student,
    StudentCourse,
    SupportRequest,
    Teacher,
    fee_type_choice,
    notice_category_choice,
    sgpa_for,
    test_name,
    time_slots,
)

User = get_user_model()
logger = logging.getLogger(__name__)

# Two weeks - long enough to be useful on a personal device, short enough that
# a shared machine does not stay signed in indefinitely.
REMEMBER_ME_SECONDS = 60 * 60 * 24 * 14

WEEKDAY_INDEX = {name: index for index, (name, _) in enumerate(DAYS_OF_WEEK)}

# Overdue is not one of Fee.status's three values - it cuts across them - but
# it is the filter staff actually reach for, so it sits in the same control.
FEE_STATUS_FILTERS = (
    ('unpaid', 'Unpaid'),
    ('partial', 'Partially paid'),
    ('paid', 'Paid'),
    ('overdue', 'Overdue'),
)

# Create your views here.


@login_required
def index(request):
    if request.user.is_teacher:
        return render(request, 'info/t_homepage.html',
                      _teacher_dashboard(request.user.teacher))
    if request.user.is_student:
        return render(request, 'info/homepage.html',
                      _student_dashboard(request.user.student))
    if request.user.is_superuser:
        return render(request, 'info/admin_page.html', _admin_dashboard())
    return render(request, 'info/logout.html')


def _student_dashboard(student):
    """Attendance standing, fees due and notices, rather than four links.

    Everything here is already computed elsewhere - classes_to_attend and
    classes_can_skip have been model properties all along and were never shown
    anywhere.
    """
    courses = Course.objects.filter(assign__class_id=student.class_id_id).distinct()
    AttendanceTotal.objects.bulk_create(
        [AttendanceTotal(student=student, course=c) for c in courses],
        ignore_conflicts=True,
    )
    totals = list(AttendanceTotal.objects
                  .filter(student=student, course__in=courses)
                  .select_related('course')
                  .with_counts())

    held = sum(t.total_class for t in totals)
    attended = sum(t.att_class for t in totals)
    fees = list(student.fees.all())

    return {
        'latest_notices': Notice.objects.filter(
            audience__in=['All', 'Students'])[:3],
        # Weighted across all sessions, not the mean of the percentages, which
        # would over-weight a course with only a handful of classes.
        'overall_attendance': round(attended / held * 100, 2) if held else None,
        'at_risk': [t for t in totals if t.has_classes and t.attendance < 75],
        'can_skip': [t for t in totals if t.has_classes and t.classes_can_skip],
        'fees_due': sum(f.balance for f in fees),
        'next_due': min((f for f in fees if f.balance > 0),
                        key=lambda f: f.due_date, default=None),
        'overdue_count': sum(1 for f in fees if f.is_overdue),
    }


def _teacher_dashboard(teacher):
    """What still needs doing, rather than a menu of sections."""
    today = timezone.localdate()
    assigns = list(Assign.objects.filter(teacher=teacher)
                   .select_related('course', 'class_id'))

    # A past session still at CLASS_PENDING is attendance the teacher owes.
    pending_sessions = (AttendanceClass.objects
                        .filter(assign__in=assigns, status=CLASS_PENDING,
                                date__lte=today)
                        .select_related('assign__course', 'assign__class_id')
                        .order_by('-date')[:10])

    # Today's unmarked sessions, so the dashboard can offer the one thing a
    # teacher opens this app to do. Four clicks - Classes, class, ClassDates,
    # pick the date - for something that is unambiguous.
    todays_sessions = (AttendanceClass.objects
                       .filter(assign__in=assigns, status=CLASS_PENDING,
                               date=today)
                       .select_related('assign__course', 'assign__class_id'))
    pending_marks = (MarksClass.objects
                     .filter(assign__in=assigns, status=False)
                     .select_related('assign__course', 'assign__class_id'))

    student_count = Student.objects.filter(
        class_id__in={a.class_id_id for a in assigns}).count()

    at_risk = [t for t in attendance_rows(
        students=Student.objects.filter(
            class_id__in={a.class_id_id for a in assigns}),
        courses=[a.course_id for a in assigns])
        if t.has_classes and t.attendance < 75]

    return {
        'latest_notices': Notice.objects.filter(
            audience__in=['All', 'Teachers'])[:3],
        'assigns': assigns,
        'class_count': len(assigns),
        'student_count': student_count,
        'pending_sessions': pending_sessions,
        'todays_sessions': todays_sessions,
        'pending_sessions_count': AttendanceClass.objects.filter(
            assign__in=assigns, status=CLASS_PENDING, date__lte=today).count(),
        'pending_marks': pending_marks,
        'at_risk': sorted(at_risk, key=lambda t: t.attendance)[:10],
        'at_risk_count': len(at_risk),
        # Which of their own classes is doing worse - a comparison the teacher
        # could previously only make by opening each report in turn.
        'attendance_by_class': _teacher_class_attendance(assigns),
    }


def _teacher_class_attendance(assigns):
    """Average attendance per class this teacher takes, weakest first.

    Keyed by assignment rather than class, since the same class can appear twice
    under two courses and they are separate attendance registers.
    """
    counts = (Attendance.objects
              .filter(course__in=[a.course_id for a in assigns],
                      student__class_id__in={a.class_id_id for a in assigns})
              .values('course', 'student__class_id')
              .annotate(held=Count('pk'),
                        attended=Count('pk', filter=Q(status=True))))
    by_pair = {(row['course'], row['student__class_id']): row for row in counts}

    out = []
    for assign in assigns:
        row = by_pair.get((assign.course_id, assign.class_id_id))
        if not row or not row['held']:
            continue
        label = '%s %s' % (assign.class_id_id, assign.course.shortname)
        out.append((label, round(row['attended'] / row['held'] * 100, 1)))
    return sorted(out, key=lambda pair: pair[1])


def _admin_dashboard():
    with_classes = [t for t in attendance_rows() if t.has_classes]
    held = sum(t.total_class for t in with_classes)
    attended = sum(t.att_class for t in with_classes)

    fees = Fee.objects.all()
    billed = sum(f.amount for f in fees)
    collected = sum(f.paid_amount for f in fees)
    outstanding = billed - collected

    return {
        'latest_notices': Notice.objects.all()[:3],
        'student_count': Student.objects.count(),
        'teacher_count': Teacher.objects.count(),
        'dept_count': Dept.objects.count(),
        'avg_attendance': round(attended / held * 100, 2) if held else None,
        'at_risk_count': len({t.student.USN for t in with_classes
                              if t.attendance < 75}),
        'fees_outstanding': outstanding,
        'overdue_count': sum(1 for f in fees if f.is_overdue),
        'open_support': SupportRequest.objects.exclude(status='Resolved').count(),
        'recent_activity': AuditLog.objects.all()[:10],
        # The dashboard reported these as bare numbers; which classes are
        # struggling and where the money is owed both need a comparison to read.
        'attendance_by_class': _attendance_by_class(with_classes),
        'collection_rate': (float(collected / billed * 100) if billed else None),
        'outstanding_by_type': _outstanding_by_fee_type(fees),
        'students_by_dept': _students_by_dept(),
    }


def _attendance_by_class(rows):
    """Average attendance per class, weakest first."""
    classes = {s.USN: s.class_id_id for s in Student.objects.only('USN', 'class_id')}
    totals = {}
    for row in rows:
        key = classes.get(row.student.USN)
        if key is None:
            continue
        held, attended = totals.setdefault(key, [0, 0])
        totals[key] = [held + row.total_class, attended + row.att_class]

    out = [(name, round(attended / held * 100, 1))
           for name, (held, attended) in totals.items() if held]
    return sorted(out, key=lambda pair: pair[1])


def _outstanding_by_fee_type(fees):
    """What is owed, by fee type - largest first, in thousands of rupees.

    Full rupee amounts run past the chart's value column and read as a wall of
    digits; the reader of a ranked comparison wants the relative sizes, and the
    exact figures are on the fees page.
    """
    totals = {}
    for fee in fees:
        if fee.balance > 0:
            totals[fee.fee_type] = totals.get(fee.fee_type, 0) + float(fee.balance)
    return sorted(((name, round(amount / 1000, 1)) for name, amount in totals.items()),
                  key=lambda pair: -pair[1])


def _students_by_dept():
    rows = (Student.objects
            .values('class_id__dept__name')
            .annotate(total=Count('USN'))
            .order_by('-total'))
    return [(row['class_id__dept__name'] or 'Unassigned', row['total'])
            for row in rows]


@login_required()
def attendance(request, stud_id):
    stud = get_object_or_404(Student, USN=stud_id)
    courses = Course.objects.filter(assign__class_id=stud.class_id_id).distinct()

    # Backfill in one go rather than one get-or-create per course inside a loop.
    AttendanceTotal.objects.bulk_create(
        [AttendanceTotal(student=stud, course=c) for c in courses],
        ignore_conflicts=True,
    )
    att_list = (AttendanceTotal.objects
                .filter(student=stud, course__in=courses)
                .select_related('course', 'student')
                .with_counts())
    return render(request, 'info/attendance.html', {'att_list': att_list})


@login_required()
def attendance_detail(request, stud_id, course_id):
    stud = get_object_or_404(Student, USN=stud_id)
    cr = get_object_or_404(Course, id=course_id)
    att_list = Attendance.objects.filter(course=cr, student=stud).order_by('date')
    return render(request, 'info/att_detail.html', {'att_list': att_list, 'cr': cr})


# Teacher Views

@login_required
@teacher_required
@owns_teacher_id('teacher_id')
def t_clas(request, teacher_id, choice):
    teacher1 = get_object_or_404(Teacher, id=teacher_id)
    return render(request, 'info/t_clas.html', {'teacher1': teacher1, 'choice': choice})


@login_required()
@teacher_required
@owns_assign('assign_id')
def t_student(request, assign_id):
    ass = get_object_or_404(Assign, id=assign_id)
    students = list(ass.class_id.student_set.all())

    AttendanceTotal.objects.bulk_create(
        [AttendanceTotal(student=s, course=ass.course) for s in students],
        ignore_conflicts=True,
    )
    att_list = (AttendanceTotal.objects
                .filter(student__in=students, course=ass.course)
                .select_related('course', 'student')
                .with_counts())
    return render(request, 'info/t_students.html', {'att_list': att_list})


@login_required()
@teacher_required
@owns_assign('assign_id')
def t_class_date(request, assign_id):
    """Every session for this class, upcoming ones included.

    This was filtered to date__lte=now, so a teacher could not see what was
    coming - and today's session only appeared once the day was already under
    way. Upcoming rows render as locked rather than markable.
    """
    ass = get_object_or_404(Assign, id=assign_id)
    att_list = ass.attendanceclass_set.order_by('-date')
    today = timezone.localdate()
    return render(request, 'info/t_class_date.html', {
        'att_list': att_list,
        'assign': ass,
        'today_session': next(
            (a for a in att_list if a.date == today
             and a.status == CLASS_PENDING), None),
    })


@login_required()
@teacher_required
@owns_attendance_class('ass_c_id')
@require_POST
def cancel_class(request, ass_c_id):
    assc = get_object_or_404(AttendanceClass, id=ass_c_id)
    assc.status = CLASS_CANCELLED
    assc.save()
    AuditLog.record(
        actor=request.user, action='attendance.cancelled', target=assc,
        summary='Class cancelled: %s on %s' % (assc.assign.class_id, assc.date))
    return HttpResponseRedirect(reverse('t_class_date', args=(assc.assign_id,)))


@login_required()
@teacher_required
@owns_attendance_class('ass_c_id')
def t_attendance(request, ass_c_id):
    assc = get_object_or_404(
        AttendanceClass.objects.select_related('assign__course',
                                               'assign__class_id'),
        id=ass_c_id)
    if not assc.is_markable:
        messages.error(request, 'That session cannot be marked: it is %s.'
                       % assc.state)
        return redirect('t_class_date', assign_id=assc.assign_id)

    ass = assc.assign
    c = ass.class_id
    return render(request, 'info/t_attendance.html', {
        'ass': ass,
        'c': c,
        'assc': assc,
        'students': c.student_set.order_by('name'),
    })


@login_required()
@teacher_required
@owns_attendance_class('ass_c_id')
def edit_att(request, ass_c_id):
    assc = get_object_or_404(AttendanceClass, id=ass_c_id)
    cr = assc.assign.course
    att_list = Attendance.objects.filter(attendanceclass=assc, course=cr)
    context = {
        'assc': assc,
        'att_list': att_list,
    }
    return render(request, 'info/t_edit_att.html', context)


@login_required()
@teacher_required
@owns_attendance_class('ass_c_id')
@require_POST
def confirm(request, ass_c_id):
    assc = get_object_or_404(AttendanceClass, id=ass_c_id)

    # The rules live in info.services because the API submits attendance too,
    # and two copies of them would drift.
    present = {usn for usn, value in request.POST.items() if value == 'present'}
    try:
        submit_attendance(assc, present, request.user)
    except SessionNotMarkable as exc:
        # The form is only reachable for a markable session, but the POST
        # endpoint has to say so itself - nothing stops a hand-built request.
        messages.error(request, str(exc))
        return redirect('t_class_date', assign_id=assc.assign_id)

    return HttpResponseRedirect(reverse('t_class_date', args=(assc.assign_id,)))


@login_required()
@teacher_required
def t_attendance_detail(request, stud_id, course_id):
    stud = get_object_or_404(Student, USN=stud_id)
    cr = get_object_or_404(Course, id=course_id)
    assert_teaches(request, cr.id, stud)
    att_list = Attendance.objects.filter(course=cr, student=stud).order_by('date')
    return render(request, 'info/t_att_detail.html', {'att_list': att_list, 'cr': cr})


@login_required()
@teacher_required
@require_POST
def change_att(request, att_id):
    # Was a GET that flipped the record, so it bypassed CSRF entirely and could
    # be fired by anything that could make the browser issue a request. It is a
    # POST now, and restricted to the teacher who takes that course.
    a = get_object_or_404(Attendance, id=att_id)
    assert_teaches(request, a.course_id, a.student)
    was = a.status
    a.status = not a.status
    a.save()
    AuditLog.record(
        actor=request.user, action='attendance.changed', target=a,
        student=a.student,
        summary='%s on %s for %s' % (
            'Marked present' if a.status else 'Marked absent',
            a.date, a.course_id),
        changes={'status': {'from': was, 'to': a.status}},
    )
    return HttpResponseRedirect(reverse('t_attendance_detail', args=(a.student.USN, a.course_id)))


@login_required()
@teacher_required
@owns_assign('assign_id')
def t_extra_class(request, assign_id):
    ass = get_object_or_404(Assign, id=assign_id)
    c = ass.class_id
    context = {
        'ass': ass,
        'c': c,
    }
    return render(request, 'info/t_extra_class.html', context)


@login_required()
@teacher_required
@owns_assign('assign_id')
@require_POST
def e_confirm(request, assign_id):
    ass = get_object_or_404(Assign, id=assign_id)
    cr = ass.course
    cl = ass.class_id

    # The date arrived straight from the form and went into the database
    # unchecked, so a typo or a hand-crafted post created sessions on any date
    # at all - including outside the semester, or a second one for a day that
    # already had a class.
    form = ExtraClassForm(request.POST, assign=ass)
    if not form.is_valid():
        return render(request, 'info/t_extra_class.html',
                      {'ass': ass, 'c': cl, 'form': form})

    session_date = form.cleaned_data['date']
    with transaction.atomic():
        assc = ass.attendanceclass_set.create(status=1, date=session_date)
        for s in cl.student_set.all():
            Attendance.objects.create(
                course=cr, student=s, attendanceclass=assc, date=session_date,
                status=request.POST.get(s.USN) == 'present',
            )

    return HttpResponseRedirect(reverse('t_clas', args=(ass.teacher_id, 1)))


@login_required()
@teacher_required
@owns_assign('assign_id')
def t_report(request, assign_id):
    ass = get_object_or_404(Assign, id=assign_id)
    students = list(ass.class_id.student_set.all())

    # A bare .get() here meant one student without a StudentCourse row took the
    # whole class report down with an uncaught DoesNotExist.
    StudentCourse.objects.bulk_create(
        [StudentCourse(student=s, course=ass.course) for s in students],
        ignore_conflicts=True,
    )
    sc_list = _report_rows(ass, students)

    order = request.GET.get('sort', 'name')
    sc_list = _sort_report(sc_list, order)
    at_risk_only = request.GET.get('risk') == '1'

    at_risk = [sc for sc in sc_list if sc.is_at_risk]
    held = [sc for sc in sc_list if sc.has_marks]

    return render(request, 'info/t_report.html', {
        'assign': ass,
        'sc_list': [sc for sc in sc_list if sc.is_at_risk] if at_risk_only else sc_list,
        'sort': order,
        'sorts': REPORT_SORT_LABELS,
        'at_risk_only': at_risk_only,
        'headcount': len(sc_list),
        'at_risk_count': len(at_risk),
        # Averages over the students who actually have something recorded, so a
        # class that has not sat a test yet does not report an average of zero.
        'avg_cie': (round(sum(sc.get_cie() for sc in held) / len(held), 1)
                    if held else None),
        'avg_attendance': (round(sum(sc.get_attendance() for sc in sc_list) / len(sc_list), 1)
                           if sc_list else None),
        'ineligible_count': sum(1 for sc in sc_list if sc.is_see_eligible is False),
        'cie_max': CIE_MAX,
        'eligibility_cie': SEE_ELIGIBILITY_CIE,
    })


def _report_rows(assign, students):
    """Every student on this class's course, with marks and attendance loaded.

    Both attach_* calls exist to keep this flat: without them each row looks up
    its own attendance and its own submitted-components status, and the
    template asks for each twice.
    """
    # attach_attendance reads AttendanceTotal, and those rows are only
    # backfilled when someone opens the attendance page - so without this the
    # report showed 0% for every student until they each did, and then flagged
    # the whole class as at risk. Same guard the attendance page already has.
    AttendanceTotal.objects.bulk_create(
        [AttendanceTotal(student=s, course=assign.course) for s in students],
        ignore_conflicts=True,
    )

    rows = list(StudentCourse.objects
                .filter(student__in=students, course=assign.course)
                .select_related('student', 'course')
                .prefetch_related('marks_set'))
    StudentCourse.attach_attendance(rows, assign.course)
    StudentCourse.attach_submitted(rows, assign.class_id_id)
    return rows


REPORT_SORTS = {
    'name': lambda sc: sc.student.name,
    'usn': lambda sc: sc.student_id,
    'cie': lambda sc: sc.get_cie(),
    'attendance': lambda sc: sc.get_attendance(),
}

REPORT_SORT_LABELS = (
    ('name', 'Name'),
    ('usn', 'USN'),
    ('cie', 'Lowest CIE'),
    ('attendance', 'Lowest attendance'),
)


def _sort_report(rows, order):
    """Lowest-first on the two numeric columns - the point of sorting a class
    report is to bring the students in trouble to the top."""
    return sorted(rows, key=REPORT_SORTS.get(order, REPORT_SORTS['name']))


@login_required()
@teacher_required
@owns_assign('assign_id')
def t_report_export(request, assign_id):
    """The same report as a spreadsheet, for department records."""
    ass = get_object_or_404(Assign, id=assign_id)
    rows = _sort_report(
        _report_rows(ass, list(ass.class_id.student_set.all())),
        request.GET.get('sort', 'name'))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Class report'
    ws.append(['USN', 'Name', 'Attendance %', 'CIE', 'SEE eligible', 'At risk'])
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4F46E5', end_color='4F46E5',
                                fill_type='solid')

    for sc in rows:
        eligible = sc.is_see_eligible
        ws.append([
            sc.student_id,
            sc.student.name,
            sc.get_attendance(),
            sc.get_cie(),
            # Undecided is its own answer here, not a False.
            'Undecided' if eligible is None else ('Yes' if eligible else 'No'),
            'Yes' if sc.is_at_risk else '',
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = (
        'attachment; filename="%s_%s_report.xlsx"'
        % (ass.class_id_id, ass.course_id))
    wb.save(response)
    return response


BREAK_COLUMNS = (4, 8)


def _timetable_matrix(slots, cell, blank=''):
    """Lay `slots` out as the 6x12 grid the timetable templates render.

    Column 0 is the day label and columns 4 and 8 are the break and lunch gaps;
    the remaining nine map to time_slots in order.

    This used to call .get() once per cell - 54 queries per page, every miss
    raising DoesNotExist as normal control flow - and caught only DoesNotExist,
    so two courses scheduled against one class in the same slot raised
    MultipleObjectsReturned and 500'd the page for everyone in that class.
    Indexing what a single query returned removes both problems.
    """
    by_slot = {(s.day, s.period): s for s in slots}

    matrix = []
    for day, _ in DAYS_OF_WEEK:
        row = [day]
        period_index = 0
        for column in range(1, 12):
            if column in BREAK_COLUMNS:
                row.append(blank)
                continue
            slot = by_slot.get((day, time_slots[period_index][0]))
            row.append(cell(slot) if slot else blank)
            period_index += 1
        matrix.append(row)
    return matrix


@login_required()
def timetable(request, class_id):
    slots = (AssignTime.objects
             .filter(assign__class_id=class_id)
             .select_related('assign'))
    matrix = _timetable_matrix(slots, cell=lambda s: s.assign.course_id)
    return render(request, 'info/timetable.html', {'matrix': matrix})


@login_required()
@teacher_required
@owns_teacher_id('teacher_id')
def t_timetable(request, teacher_id):
    slots = (AssignTime.objects
             .filter(assign__teacher_id=teacher_id)
             .select_related('assign', 'assign__course', 'assign__class_id'))
    # The template checks `j == True` for an empty cell, which is why the blank
    # here is True rather than '' as in the student view.
    class_matrix = _timetable_matrix(slots, cell=lambda s: s, blank=True)
    return render(request, 'info/t_timetable.html', {'class_matrix': class_matrix})


def _next_weekday(day, today=None):
    """The next date `day` falls on, counting today as the next occurrence.

    A timetable slot is a recurring weekday, but a cancellation belongs to one
    date - so answering "is this teacher free" needs a date to look at.
    Returns None for a day name that is not in DAYS_OF_WEEK, in which case
    callers fall back to the static timetable.
    """
    target = WEEKDAY_INDEX.get(day)
    if target is None:
        return None
    today = today or timezone.localdate()
    return today + timedelta(days=(target - today.weekday()) % 7)


@login_required()
@teacher_required
def free_teachers(request, asst_id):
    """Who across the college could cover this slot.

    The candidate pool used to be Teacher.objects.filter(assign__class_id=...),
    i.e. only teachers already teaching this class - so the page answered a
    much narrower question than its title and could never find an outside
    substitute. It also joined across `assign`, which repeated any teacher
    holding two courses for the class, and then ran one query per candidate to
    decide availability in Python.
    """
    asst = get_object_or_404(
        AssignTime.objects.select_related('assign__class_id', 'assign__course',
                                          'assign__teacher'),
        id=asst_id)

    date = _next_weekday(asst.day)

    # A cancelled session frees its teacher up even though the timetable still
    # shows them as teaching - which is the case a substitute is being looked
    # for in the first place.
    cancelled_assigns = []
    if date is not None:
        cancelled_assigns = list(
            AttendanceClass.objects
            .filter(date=date, status=CLASS_CANCELLED)
            .values_list('assign_id', flat=True))

    busy = (AssignTime.objects
            .filter(day=asst.day, period=asst.period)
            .exclude(assign_id__in=cancelled_assigns)
            .values_list('assign__teacher_id', flat=True))

    # exclude() against the collected ids rather than filtering across the
    # assign join, so a teacher cannot come back twice.
    ft_list = (Teacher.objects
               .exclude(id__in=busy)
               .select_related('dept')
               .order_by('name'))

    return render(request, 'info/free_teachers.html', {
        'ft_list': ft_list,
        'slot': asst,
        'date': date,
    })


# student marks


def _student_marks_rows(stud):
    """The student's courses with marks, visibility and rank loaded.

    Shared with the PDF marks card so the paper and the screen cannot disagree
    about what a student scored.
    """
    courses = Course.objects.filter(assign__class_id=stud.class_id_id).distinct()

    # The old fallback here passed type='I' to marks_set.create(), but Marks has
    # no `type` field - it raised TypeError. It only ran when a StudentCourse
    # row was missing, which the signals normally prevent, so the page worked by
    # luck. bulk_create with ignore_conflicts covers the same gap without the
    # per-course round trips.
    StudentCourse.objects.bulk_create(
        [StudentCourse(student=stud, course=c) for c in courses],
        ignore_conflicts=True,
    )
    sc_list = list(StudentCourse.objects
                   .filter(student=stud, course__in=courses)
                   .select_related('course')
                   .prefetch_related('marks_set'))

    # Published, not merely entered: a student sees a mark when the result is
    # released, not the instant the teacher types it.
    StudentCourse.attach_submitted(sc_list, stud.class_id_id,
                                   published_only=True)
    StudentCourse.attach_rank(
        sc_list, stud.class_id_id,
        {sc.course_id: sc.visible_components for sc in sc_list})
    return sc_list


@login_required()
def marks_list(request, stud_id):
    stud = get_object_or_404(Student, USN=stud_id)
    sc_list = _student_marks_rows(stud)

    # Courses with something actually marked. Filtering on a truthy CIE instead
    # would silently drop a course genuinely sitting at zero - which is exactly
    # the one worth calling out - while keeping one that has simply not started.
    with_cie = [sc for sc in sc_list if sc.has_marks]
    return render(request, 'info/marks_list.html', {
        'student': stud,
        'sc_list': sc_list,
        'sgpa': sgpa_for(sc_list),
        'cie_max': CIE_MAX,
        'see_max': SEE_MAX,
        'eligibility_cie': SEE_ELIGIBILITY_CIE,
        # Honest label while no result is in: an average CIE is not a GPA.
        'average_cie': (round(sum(sc.get_cie() for sc in with_cie) / len(with_cie), 1)
                        if with_cie else None),
        'best': max(with_cie, key=lambda sc: sc.get_cie(), default=None),
        'weakest': min(with_cie, key=lambda sc: sc.get_cie(), default=None),
        # Where the subjects sit relative to each other. The accordion below
        # answers one course at a time; ranking them is a different question.
        'cie_by_course': sorted(
            ((sc.course.shortname or sc.course.name, sc.get_cie())
             for sc in with_cie),
            key=lambda pair: -pair[1]),
    })


@login_required()
def marks_card(request, stud_id):
    """The marks page as a one-page PDF.

    Same access rule as the marks page itself, and the same numbers: the rows
    are built by the same helper, so the paper and the screen cannot disagree.
    """
    stud = get_object_or_404(Student, USN=stud_id)
    if not (request.user.is_superuser
            or (request.user.is_student
                and request.user.student.USN == stud.USN)):
        raise PermissionDenied

    sc_list = _student_marks_rows(stud)
    buffer = report_card(stud, sc_list, sgpa_for(sc_list), timezone.localdate())

    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = (
        'attachment; filename="%s_marks_card.pdf"' % stud.USN)
    return response


# teacher marks


@login_required()
@teacher_required
@owns_assign('assign_id')
def t_marks_list(request, assign_id):
    ass = get_object_or_404(Assign, id=assign_id)
    return render(request, 'info/t_marks_list.html', {
        'assign': ass,
        'm_list': MarksClass.objects.filter(assign=ass),
    })


@login_required()
@teacher_required
@owns_marks_class('marks_c_id')
@require_POST
def publish_marks(request, marks_c_id):
    """Release an entered batch to the students, or pull it back.

    Entry and publication were the same act, so a mark was visible the instant
    it was typed - including a slip the teacher was about to correct.
    """
    mc = get_object_or_404(MarksClass, id=marks_c_id)
    if not mc.status:
        messages.error(request, 'Enter the marks before publishing them.')
        return redirect('t_marks_list', assign_id=mc.assign_id)

    if request.POST.get('action') == 'unpublish':
        mc.unpublish()
        AuditLog.record(
            actor=request.user, action='marks.unpublished', target=mc,
            summary='%s withdrawn for %s' % (mc.name, mc.assign.class_id))
        messages.success(request, '%s is hidden from students again.' % mc.name)
    else:
        mc.publish()
        AuditLog.record(
            actor=request.user, action='marks.published', target=mc,
            summary='%s published for %s' % (mc.name, mc.assign.class_id))
        messages.success(request, '%s is now visible to students.' % mc.name)

    return redirect('t_marks_list', assign_id=mc.assign_id)


@login_required()
@teacher_required
@owns_marks_class('marks_c_id')
def t_marks_entry(request, marks_c_id):
    mc = get_object_or_404(MarksClass, id=marks_c_id)
    return render(request, 'info/t_marks_entry.html',
                  _marks_entry_context(mc, _roster(mc, request)))


def _roster(mc, request):
    order = 'name' if request.GET.get('sort') != 'usn' else 'USN'
    return list(mc.assign.class_id.student_set.order_by(order))


def _previous_component(name):
    """The component entered before this one, for context while typing.

    Seeing Internal test 1 beside Internal test 2 is how a transposed row gets
    spotted at entry time rather than after the grades are published.
    """
    names = [choice for choice, _ in test_name]
    index = names.index(name) if name in names else 0
    return names[index - 1] if index > 0 else None


def _marks_entry_context(mc, students, form=None, existing=None):
    """Build the per-student rows the marks entry template renders.

    Templates can't index a dict by a variable key, so the pairing of student to
    input value, previous mark and error list is done here.
    """
    existing = existing or {}
    previous_name = _previous_component(mc.name)
    previous = {}
    if previous_name:
        previous = {
            m.studentcourse.student_id: m.marks1
            for m in Marks.objects
            .filter(studentcourse__course=mc.assign.course,
                    studentcourse__student__in=students, name=previous_name)
            .select_related('studentcourse')
        }

    rows = []
    for s in students:
        if form is not None:
            value = form.data.get(s.USN, '')
            absent = bool(form.data.get(MarksEntryForm.absent_field(s)))
        else:
            # Blank, not 0. A pre-filled zero is indistinguishable from a mark
            # of zero, so a student the teacher scrolled past used to be
            # recorded as having scored nothing.
            value, absent = existing.get(s.USN, ('', False))
        rows.append({
            'student': s,
            'value': value,
            'absent': absent,
            'absent_field': MarksEntryForm.absent_field(s),
            'previous': previous.get(s.USN),
            'errors': form.errors_for(s) if form else [],
        })

    return {'ass': mc.assign, 'c': mc.assign.class_id, 'mc': mc,
            'rows': rows, 'form': form,
            'previous_name': previous_name,
            'is_revision': mc.status}


@login_required()
@teacher_required
@owns_marks_class('marks_c_id')
@require_POST
def marks_confirm(request, marks_c_id):
    mc = get_object_or_404(MarksClass, id=marks_c_id)
    ass = mc.assign
    # Same ordering as the form that posted, so a re-render after a validation
    # error puts the rows back where the teacher left them.
    students = _roster(mc, request)

    # Marks went in as raw POST strings with no bounds check, so a slip of the
    # keyboard stored 85 on a test worth 20 - the field validators never run on
    # a plain .save(). Validate the whole class before writing any of it.
    form = MarksEntryForm(request.POST, students=students,
                          total_marks=mc.total_marks)
    if not form.is_valid():
        return render(request, 'info/t_marks_entry.html',
                      _marks_entry_context(mc, students, form))

    # The rules live in info.services because the API enters marks too, and
    # two copies of them would drift.
    submit_marks(mc, {s: form.marks_for(s) for s in students}, request.user)

    return HttpResponseRedirect(reverse('t_marks_list', args=(ass.id,)))


@login_required()
@teacher_required
@owns_marks_class('marks_c_id')
def edit_marks(request, marks_c_id):
    """Re-open a submitted batch, on the same form that entered it.

    This used to be a second, near-identical template that had no error
    display at all - so a failed edit re-rendered the *other* template - and it
    walked the roster with bare StudentCourse.objects.get() and marks_set.get()
    calls, either of which raised DoesNotExist and 500'd the page for the whole
    class if a single row was missing.
    """
    mc = get_object_or_404(MarksClass, id=marks_c_id)
    students = _roster(mc, request)
    existing = {
        m.studentcourse.student_id: ('' if m.is_absent else m.marks1, m.is_absent)
        for m in Marks.objects
        .filter(studentcourse__course=mc.assign.course,
                studentcourse__student__in=students, name=mc.name)
        .select_related('studentcourse')
    }
    return render(request, 'info/t_marks_entry.html',
                  _marks_entry_context(mc, students, existing=existing))


@login_required()
@teacher_required
@owns_assign('assign_id')
def student_marks(request, assign_id):
    """The class roster with marks, plus what they add up to.

    A bare .get() here raised an uncaught DoesNotExist on an unknown id and
    500'd where every neighbouring view returns 404.
    """
    ass = get_object_or_404(
        Assign.objects.select_related('course', 'class_id'), id=assign_id)
    students = list(ass.class_id.student_set.order_by('name'))

    # attach_attendance reads AttendanceTotal, and those rows only appear when
    # someone opens the attendance page - without this the whole class reads as
    # 0% and every student is flagged at risk. Third page to stand in this
    # trap; the dashboards avoid it by computing from Attendance directly.
    AttendanceTotal.objects.bulk_create(
        [AttendanceTotal(student=s, course=ass.course) for s in students],
        ignore_conflicts=True,
    )

    sc_list = list(StudentCourse.objects
                   .filter(student__in=students, course=ass.course)
                   .select_related('student', 'course')
                   .prefetch_related('marks_set'))
    StudentCourse.attach_submitted(sc_list, ass.class_id_id)
    StudentCourse.attach_attendance(sc_list, ass.course)

    marked = [sc for sc in sc_list if sc.has_marks]
    cies = sorted(sc.get_cie() for sc in marked)

    return render(request, 'info/t_student_marks.html', {
        'assign': ass,
        'sc_list': sc_list,
        'cie_max': CIE_MAX,
        'headcount': len(sc_list),
        'marked_count': len(marked),
        'average_cie': round(sum(cies) / len(cies), 1) if cies else None,
        'median_cie': _median(cies),
        'lowest_cie': cies[0] if cies else None,
        'highest_cie': cies[-1] if cies else None,
        # Only students whose CIE is complete have an eligibility answer at
        # all. Counting the undecided ones as zero reads as "nobody qualifies"
        # when the truth is "not settled yet".
        'decided': [sc for sc in marked if sc.cie_is_final],
        'eligible_count': sum(1 for sc in marked if sc.is_see_eligible is True),
        'at_risk': [sc for sc in marked if sc.is_at_risk],
        'distribution': _cie_bands(cies),
    })


def _median(values):
    if not values:
        return None
    middle = len(values) // 2
    if len(values) % 2:
        return values[middle]
    return round((values[middle - 1] + values[middle]) / 2, 1)


# Ten-mark bands across the CIE's 0-50. Wide enough that a class of forty does
# not scatter into a flat line, narrow enough to show a shape.
CIE_BANDS = ((0, 10), (11, 20), (21, 30), (31, 40), (41, 50))


def _cie_bands(cies):
    """(label, count) per band - the shape of the class's marks.

    Tells the teacher whether the paper was too hard or too easy, which a
    column of per-student numbers never does.
    """
    if not cies:
        return []
    return [('%d-%d' % (low, high),
             sum(1 for cie in cies if low <= cie <= high))
            for low, high in CIE_BANDS]


@login_required()
def add_teacher(request):
    if not request.user.is_superuser:
        return redirect("/")

    credentials = None
    if request.method == 'POST':
        form = TeacherForm(request.POST)
        if form.is_valid():
            # Both writes go together: without this, a failure on the second
            # left a usable login account attached to no teacher record.
            with transaction.atomic():
                user, password = form.create_user()
                teacher = form.save(commit=False)
                teacher.user = user
                teacher.save()
            # Shown once, here only - the password is random and not recoverable.
            credentials = {'username': user.username, 'password': password,
                           'name': teacher.name}
            form = TeacherForm()
    else:
        form = TeacherForm()

    return render(request, 'info/add_teacher.html',
                  {'form': form, 'credentials': credentials})


@login_required()
def add_student(request):
    # If the user is not admin, they will be redirected to home
    if not request.user.is_superuser:
        return redirect("/")

    credentials = None
    if request.method == 'POST':
        form = StudentForm(request.POST)
        if form.is_valid():
            # See add_teacher - the two writes must not be able to come apart.
            with transaction.atomic():
                user, password = form.create_user()
                student = form.save(commit=False)
                student.user = user
                student.save()
            credentials = {'username': user.username, 'password': password,
                           'name': student.name}
            form = StudentForm()
    else:
        form = StudentForm()

    return render(request, 'info/add_student.html',
                  {'form': form, 'credentials': credentials})


# ---------------------------------------------------------------------------
# Fees
# ---------------------------------------------------------------------------

@login_required()
def fees(request, stud_id):
    stud = get_object_or_404(Student, USN=stud_id)
    if not (request.user.is_superuser or (request.user.is_student and request.user.student.USN == stud.USN)):
        return redirect('/')

    # The transaction model landed without the page that shows it, so a student
    # could see a balance drop but never when or how the money was recorded.
    fee_list = stud.fees.prefetch_related('transactions')
    total_amount = sum(f.amount for f in fee_list)
    total_paid = sum(f.paid_amount for f in fee_list)
    context = {
        'stud': stud,
        'fee_list': fee_list,
        'total_amount': total_amount,
        'total_paid': total_paid,
        'total_balance': total_amount - total_paid,
    }
    return render(request, 'info/fees.html', context)


@login_required()
def fee_receipt(request, transaction_id):
    """A PDF receipt for one payment.

    Same access rule as the fee page the payment appears on: the student it
    belongs to, or an administrator.
    """
    transaction = get_object_or_404(
        FeeTransaction.objects.select_related('fee__student', 'received_by'),
        id=transaction_id)
    student = transaction.fee.student
    if not (request.user.is_superuser
            or (request.user.is_student
                and request.user.student.USN == student.USN)):
        raise PermissionDenied

    buffer = payment_receipt(transaction, timezone.localdate())
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = (
        'attachment; filename="%s.pdf"' % transaction.receipt_no)
    return response


@login_required()
def fees_export(request, stud_id):
    stud = get_object_or_404(Student, USN=stud_id)
    if not (request.user.is_superuser or (request.user.is_student and request.user.student.USN == stud.USN)):
        return redirect('/')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Fees'
    headers = ['Fee Type', 'Description', 'Amount', 'Paid', 'Balance', 'Status', 'Due Date']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')

    for f in stud.fees.all():
        ws.append([f.fee_type, f.description, float(f.amount), float(f.paid_amount),
                   float(f.balance), f.status, f.due_date.strftime('%Y-%m-%d')])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="%s_fees.xlsx"' % stud.USN
    wb.save(response)
    return response


@login_required()
def t_fees(request):
    if not (request.user.is_superuser or request.user.is_teacher):
        return redirect('/')

    q = request.GET.get('q', '').strip()
    status = request.GET.get('status', '')
    fee_type = request.GET.get('fee_type', '')
    class_id = request.GET.get('class_id', '')

    # Class.__str__ reads dept.name, so the dept has to come along or the
    # list costs a query per row.
    fee_list = Fee.objects.select_related('student', 'student__class_id',
                                          'student__class_id__dept')
    if q:
        fee_list = fee_list.filter(Q(student__name__icontains=q)
                                   | Q(student__USN__icontains=q))
    if fee_type:
        fee_list = fee_list.filter(fee_type=fee_type)
    if class_id:
        fee_list = fee_list.filter(student__class_id=class_id)

    # Applied in the database rather than by walking rows and reading the
    # status property, which is what made this page unusable at real volume.
    if status == 'paid':
        fee_list = fee_list.paid()
    elif status == 'unpaid':
        fee_list = fee_list.unpaid()
    elif status == 'partial':
        fee_list = fee_list.partial()
    elif status == 'overdue':
        fee_list = fee_list.overdue()

    # Totals cover the whole filtered set, not just the page being looked at -
    # a summary that changed as you paged would be worse than none.
    totals = fee_list.totals()

    return render(request, 'info/t_fees.html', {
        'page': Paginator(fee_list.with_balance(), 25).get_page(
            request.GET.get('page')),
        'q': q,
        'status': status,
        'fee_type': fee_type,
        'class_id': class_id,
        'fee_types': fee_type_choice,
        'status_options': FEE_STATUS_FILTERS,
        'classes': Class.objects.select_related('dept'),
        'totals': totals,
        'overdue_count': fee_list.overdue().count(),
        # Everything except `page`, so the pager keeps the current filters.
        'querystring': urlencode({k: v for k, v in request.GET.items()
                                  if k != 'page' and v}),
    })


@login_required()
def add_fee(request):
    if not (request.user.is_superuser or request.user.is_teacher):
        return redirect('/')

    if request.method == 'POST':
        form = FeeForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('t_fees')
    else:
        form = FeeForm()

    return render(request, 'info/add_fee.html', {'form': form})


@login_required()
def add_class_fee(request):
    """Raise one fee against a whole class.

    Sixty identical form submissions for a semester exam fee was the most
    obviously missing staff action in this module.
    """
    if not (request.user.is_superuser or request.user.is_teacher):
        return redirect('/')

    if request.method == 'POST':
        form = BulkFeeForm(request.POST)
        if form.is_valid():
            fees, skipped = form.build()
            # One statement, and re-running the same assignment cannot double a
            # class's fees.
            Fee.objects.bulk_create(fees)
            if fees:
                messages.success(
                    request, 'Raised %s for %d student%s.'
                    % (form.cleaned_data['fee_type'], len(fees),
                       '' if len(fees) == 1 else 's'))
            if skipped:
                messages.info(
                    request, '%d student%s already had this fee and %s skipped.'
                    % (skipped, '' if skipped == 1 else 's',
                       'was' if skipped == 1 else 'were'))
            return redirect('t_fees')
    else:
        form = BulkFeeForm()

    return render(request, 'info/add_class_fee.html', {'form': form})


@login_required()
def edit_fee(request, fee_id):
    """Record a payment against a fee.

    This used to overwrite paid_amount with a new running total, so staff had to
    do the arithmetic themselves, a mistake was unrecoverable, and nothing
    recorded when the money came in or who took it.
    """
    if not (request.user.is_superuser or request.user.is_teacher):
        return redirect('/')

    fee = get_object_or_404(Fee.objects.select_related('student'), id=fee_id)

    if request.method == 'POST':
        form = FeeTransactionForm(request.POST, fee=fee)
        if form.is_valid():
            with transaction.atomic():
                payment = form.save(commit=False)
                payment.fee = fee
                payment.received_by = request.user
                payment.save()
                fee.recalculate_paid()
                AuditLog.record(
                    actor=request.user, action='fee.payment', target=payment,
                    student=fee.student,
                    summary='%s of %s for %s (%s)'
                            % (payment.receipt_no, payment.amount,
                               fee.fee_type, payment.mode),
                    changes={'paid_amount': {'to': str(fee.paid_amount)}},
                )
            return redirect('edit_fee', fee_id=fee.id)
    else:
        form = FeeTransactionForm(fee=fee,
                                  initial={'amount': fee.balance or None})

    return render(request, 'info/edit_fee.html', {
        'fee': fee,
        'form': form,
        'payments': fee.transactions.select_related('received_by'),
    })


# ---------------------------------------------------------------------------
# Notice board
# ---------------------------------------------------------------------------

@login_required()
def notices(request):
    """The board, with search, filters and pagination.

    Previously every notice for the audience came back in one unpaginated list
    with no way to find anything in it.
    """
    notice_list = Notice.objects.visible_to(request.user).select_related('posted_by')

    query = request.GET.get('q', '').strip()
    if query:
        notice_list = notice_list.filter(
            Q(title__icontains=query) | Q(message__icontains=query))

    category = request.GET.get('category', '')
    if category:
        notice_list = notice_list.filter(category=category)

    period = request.GET.get('period', '')
    if period == 'week':
        notice_list = notice_list.filter(
            created_at__gte=timezone.now() - timedelta(days=7))
    elif period == 'month':
        notice_list = notice_list.filter(
            created_at__gte=timezone.now() - timedelta(days=30))

    page = Paginator(notice_list, 10).get_page(request.GET.get('page'))
    read_ids = set(NoticeRead.objects
                   .filter(user=request.user, notice__in=page.object_list)
                   .values_list('notice_id', flat=True))

    return render(request, 'info/notices.html', {
        'page': page,
        'q': query,
        'category': category,
        'period': period,
        'categories': notice_category_choice,
        'read_ids': read_ids,
        'unread_count': unread_notice_count(request.user),
        'can_post': request.user.is_superuser or request.user.is_teacher,
    })


@login_required()
def notice_detail(request, notice_id):
    notice = get_object_or_404(
        Notice.objects.visible_to(request.user).select_related('posted_by'),
        id=notice_id)

    # Opening it is what counts as reading it.
    NoticeRead.objects.get_or_create(notice=notice, user=request.user)

    return render(request, 'info/notice_detail.html', {
        'notice': notice,
        'can_edit': request.user.is_superuser or notice.posted_by_id == request.user.id,
        'read_count': notice.reads.count(),
    })


@login_required()
def add_notice(request):
    if not (request.user.is_superuser or request.user.is_teacher):
        return redirect('/')

    if request.method == 'POST':
        form = NoticeForm(request.POST, user=request.user)
        if form.is_valid():
            notice = form.save(commit=False)
            notice.posted_by = request.user
            notice.save()
            return redirect('notice_detail', notice_id=notice.id)
    else:
        form = NoticeForm(user=request.user)

    return render(request, 'info/add_notice.html', {'form': form})


@login_required()
def edit_notice(request, notice_id):
    """Neither editing nor deleting existed - a notice, once posted, was
    permanent from the UI."""
    notice = get_object_or_404(Notice, id=notice_id)
    if not (request.user.is_superuser or notice.posted_by_id == request.user.id):
        raise PermissionDenied

    if request.method == 'POST':
        form = NoticeForm(request.POST, instance=notice, user=request.user)
        if form.is_valid():
            form.save()
            return redirect('notice_detail', notice_id=notice.id)
    else:
        form = NoticeForm(instance=notice, user=request.user)

    return render(request, 'info/add_notice.html',
                  {'form': form, 'notice': notice})


@login_required()
@require_POST
def delete_notice(request, notice_id):
    notice = get_object_or_404(Notice, id=notice_id)
    if not (request.user.is_superuser or notice.posted_by_id == request.user.id):
        raise PermissionDenied

    notice.delete()
    messages.success(request, 'Notice deleted.')
    return redirect('notices')


def unread_notice_count(user):
    """Notices the user has not opened, for the topbar badge."""
    if not user.is_authenticated:
        return 0
    return (Notice.objects.visible_to(user)
            .exclude(reads__user=user)
            .count())

class ErpPasswordChangeView(PasswordChangeView):
    """Django's password change, plus clearing the must-change flag.

    Without this the middleware would keep redirecting a user who had just
    picked a new password straight back to this page.
    """
    template_name = 'info/password_change.html'
    success_url = reverse_lazy('password_change_done')

    def form_valid(self, form):
        response = super().form_valid(form)
        if self.request.user.must_change_password:
            self.request.user.must_change_password = False
            self.request.user.save(update_fields=['must_change_password'])
        return response


class ErpLoginView(LoginView):
    """Login with a role selector, Remember Me and the support form alongside."""
    template_name = 'info/login.html'
    authentication_form = ErpLoginForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.setdefault('support_form', SupportRequestForm())
        return context

    def form_valid(self, form):
        # Without this every session lasts SESSION_COOKIE_AGE regardless, so the
        # checkbox would be decoration.
        if form.cleaned_data.get('remember_me'):
            self.request.session.set_expiry(REMEMBER_ME_SECONDS)
        else:
            self.request.session.set_expiry(0)  # ends when the browser closes
        return super().form_valid(form)


def support_request(request):
    """Handle the "Contact Administrator" form on the login page.

    Deliberately open to anonymous callers - someone who cannot sign in is
    exactly who needs it - so it is rate-limited per IP and the form carries a
    honeypot.
    """
    if request.method != 'POST':
        return redirect('login')

    form = SupportRequestForm(request.POST)
    if _support_rate_limited(request):
        form.add_error(None, 'Too many messages from here just now. '
                             'Try again in a few minutes.')
    elif form.is_valid():
        support = form.save()
        logger.info('Support request %s from %s (%s)',
                    support.pk, support.name, support.category)
        messages.success(
            request,
            'Thanks - your message is with the administrator. '
            'They will get back to you by email.')
        return redirect('login')

    # Re-render the login page with the modal's errors, so nothing is retyped.
    return ErpLoginView.as_view(extra_context={
        'support_form': form, 'open_support': True})(request)


def _support_rate_limited(request, limit=3, window=900):
    """Allow `limit` submissions per IP per `window` seconds.

    Uses the cache rather than a table - this only needs to make spamming
    inconvenient, and losing the counter on a restart is not a problem.
    """
    ip = (request.META.get('HTTP_X_FORWARDED_FOR', '').split(',')[0].strip()
          or request.META.get('REMOTE_ADDR', 'unknown'))
    key = 'support-requests:%s' % ip
    seen = cache.get(key, 0)
    if seen >= limit:
        return True
    cache.set(key, seen + 1, window)
    return False


@login_required
def profile(request):
    """Your own details, and the contact fields you may change.

    Nobody could see or edit any of this; the only self-service the app had was
    the password form.
    """
    person = (request.user.student if request.user.is_student
              else request.user.teacher if request.user.is_teacher
              else None)
    if person is None:
        # Superusers have no Student or Teacher record to show.
        return redirect('password_change')

    if request.method == 'POST':
        form = ProfileForm(request.POST, request.FILES, profile=person,
                           user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Your details have been updated.')
            return redirect('profile')
    else:
        form = ProfileForm(profile=person, user=request.user)

    return render(request, 'info/profile.html', {
        'person': person,
        'form': form,
        'is_student': request.user.is_student,
        # Successful sign-ins only. A list of failed attempts on your own
        # account is alarming without being actionable, and the ones worth
        # investigating are visible to an administrator.
        'logins': request.user.login_events.filter(successful=True)[:10],
    })


@login_required
@teacher_required
def directory(request):
    """Look somebody up by name, USN or staff id.

    There was no directory at all - the only way to find a student was the
    search box on the fees page, and teachers were not listed anywhere.
    """
    query = request.GET.get('q', '').strip()
    kind = request.GET.get('kind', 'students')
    # Deactivated people keep their records but drop out of the working list -
    # a directory of every student who ever enrolled is not a directory.
    show_inactive = request.GET.get('inactive') == '1'

    if kind == 'teachers':
        people = Teacher.objects.select_related('dept', 'user').order_by('name')
        if query:
            people = people.filter(Q(name__icontains=query) |
                                   Q(id__icontains=query) |
                                   Q(dept__name__icontains=query))
    else:
        kind = 'students'
        people = (Student.objects
                  .select_related('class_id', 'class_id__dept', 'user')
                  .order_by('class_id_id', 'name'))
        if query:
            people = people.filter(Q(name__icontains=query) |
                                   Q(USN__icontains=query) |
                                   Q(class_id__id__icontains=query))

    if not show_inactive:
        people = people.filter(is_active=True)

    return render(request, 'info/directory.html', {
        'page': Paginator(people, 25).get_page(request.GET.get('page')),
        'q': query,
        'kind': kind,
        'show_inactive': show_inactive,
        'total': people.count(),
    })


@login_required
def bulk_import(request):
    """Enrol a whole intake from a spreadsheet.

    Two steps on purpose: validate and show what is wrong, then commit. A file
    that silently imported the rows it liked and dropped the rest would be
    worse than one that refused.
    """
    if not request.user.is_superuser:
        raise PermissionDenied

    kind = request.GET.get('kind') or request.POST.get('kind') or 'students'
    if kind not in ('students', 'teachers'):
        kind = 'students'

    columns = (STUDENT_COLUMNS if kind == 'students' else TEACHER_COLUMNS)
    context = {'kind': kind, 'columns': columns, 'max_rows': MAX_ROWS}

    if request.method == 'POST' and request.FILES.get('file'):
        try:
            rows = read_rows(request.FILES['file'], columns)
        except ImportError_ as exc:
            context['file_error'] = str(exc)
            return render(request, 'info/bulk_import.html', context)

        forms_, errors = validate(rows, kind)
        context.update({'row_count': len(rows), 'errors': errors})

        if errors:
            return render(request, 'info/bulk_import.html', context)

        created = commit(forms_, kind)
        AuditLog.record(
            actor=request.user, action='accounts.imported', target=None,
            summary='Imported %d %s from a spreadsheet' % (len(created), kind))
        # Shown once. The passwords are random and are not recoverable
        # afterwards, so this page is the only chance to hand them out.
        context['created'] = created
        return render(request, 'info/bulk_import.html', context)

    return render(request, 'info/bulk_import.html', context)


# Keys the reset flow keeps in the session between its three screens. Only the
# user id matters, and it is only set once a code has actually been issued.
RESET_USER_KEY = 'password_reset_user'
RESET_VERIFIED_KEY = 'password_reset_verified'


def _send_reset_code(user, code):
    subject = 'Your College ERP password reset code'
    body = (
        'Hello %s,\n\n'
        'Your password reset code is %s.\n\n'
        'It expires in 10 minutes and can be used once. If you did not ask to '
        'reset your password, you can ignore this - your current password '
        'still works.\n'
    ) % (user.get_username(), code)
    send_mail(subject, body, settings.DEFAULT_FROM_EMAIL, [user.email],
              fail_silently=False)


def password_reset_request(request):
    """Step one: take a username or email and, if it matches, send a code."""
    if request.method == 'POST':
        form = PasswordResetRequestForm(request.POST)
        if form.is_valid():
            user = form.find_user()

            # The response is the same whether or not the account exists, and
            # whether or not it is rate limited. Saying "no such user" here
            # turns this page into a way to find out which usernames are real.
            if user and user.email and not PasswordResetOTP.rate_limited(user):
                _, code = PasswordResetOTP.issue(user)
                try:
                    _send_reset_code(user, code)
                except Exception:
                    # A mail failure must not tell the caller they guessed a
                    # real account, so it is logged and swallowed.
                    logger.exception('Could not send a reset code to user %s',
                                     user.pk)
                else:
                    logger.info('Issued a password reset code to user %s', user.pk)

            if user:
                request.session[RESET_USER_KEY] = user.pk
            else:
                # Clearing rather than leaving a stale id behind, so a second
                # attempt cannot inherit the first one's account.
                request.session.pop(RESET_USER_KEY, None)
            request.session.pop(RESET_VERIFIED_KEY, None)

            messages.info(
                request,
                'If that account exists, a six-digit code is on its way to the '
                'email address registered against it.')
            return redirect('password_reset_verify')
    else:
        form = PasswordResetRequestForm()

    return render(request, 'info/password_reset_request.html', {'form': form})


def password_reset_verify(request):
    """Step two: check the code."""
    user_id = request.session.get(RESET_USER_KEY)

    if request.method == 'POST':
        form = PasswordResetVerifyForm(request.POST)
        if form.is_valid():
            otp = None
            if user_id:
                otp = (PasswordResetOTP.objects
                       .filter(user_id=user_id).live().first())

            if otp and otp.verify(form.cleaned_data['code']):
                request.session[RESET_VERIFIED_KEY] = otp.pk
                logger.info('Password reset code verified for user %s', user_id)
                return redirect('password_reset_set')

            # One message for a wrong code, an expired one, a locked one and an
            # identifier that never matched anything.
            form.add_error('code', 'That code is not valid. It may have '
                                   'expired, or been used already.')
    else:
        form = PasswordResetVerifyForm()

    return render(request, 'info/password_reset_verify.html', {'form': form})


def password_reset_set(request):
    """Step three: choose a new password."""
    otp_id = request.session.get(RESET_VERIFIED_KEY)
    user_id = request.session.get(RESET_USER_KEY)
    # Reached only by having verified a code in this session; the check is
    # repeated here because the URL is guessable.
    otp = PasswordResetOTP.objects.filter(pk=otp_id, user_id=user_id).first()
    if otp is None:
        return redirect('password_reset')

    if request.method == 'POST':
        form = SetPasswordForm(otp.user, request.POST)
        if form.is_valid():
            with transaction.atomic():
                form.save()
                # Whatever else was outstanding is void now, and the flag that
                # forces a change is satisfied by having just made one.
                PasswordResetOTP.objects.filter(user=otp.user).live().update(
                    used_at=timezone.now())
                if otp.user.must_change_password:
                    otp.user.must_change_password = False
                    otp.user.save(update_fields=['must_change_password'])

            request.session.pop(RESET_USER_KEY, None)
            request.session.pop(RESET_VERIFIED_KEY, None)
            logger.info('Password reset completed for user %s', otp.user_id)
            messages.success(request,
                             'Your password has been changed. Sign in with it.')
            return redirect('login')
    else:
        form = SetPasswordForm(otp.user)

    return render(request, 'info/password_reset_set.html', {'form': form})
