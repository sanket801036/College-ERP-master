"""Bulk enrolment from a spreadsheet.

Adding a sixty-student intake meant sixty trips through the add-student form.

The rows go through StudentForm and TeacherForm, the same forms the web pages
use, rather than straight into the database. A bulk path with its own
validation would be a way around the checks those forms perform - the duplicate
primary key that silently overwrote an existing record, most of all.
"""
import csv
import io

from django.db import transaction
from openpyxl import load_workbook

from info.forms import StudentForm, TeacherForm

STUDENT_COLUMNS = ('usn', 'name', 'class', 'sex', 'dob', 'email')
TEACHER_COLUMNS = ('staff_id', 'name', 'department', 'sex', 'dob', 'email')

# Header -> form field. The spreadsheet uses the words an administrator would
# write; the form uses the model's names.
STUDENT_FIELDS = {'usn': 'USN', 'name': 'name', 'class': 'class_id',
                  'sex': 'sex', 'dob': 'DOB', 'email': 'email'}
TEACHER_FIELDS = {'staff_id': 'id', 'name': 'name', 'department': 'dept',
                  'sex': 'sex', 'dob': 'DOB', 'email': 'email'}

MAX_ROWS = 500


class ImportError_(Exception):
    """The file itself is unusable - wrong type, no header, far too big."""


def _normalise(value):
    if value is None:
        return ''
    # openpyxl hands back dates and numbers as objects; the forms want the text
    # a person typed.
    if hasattr(value, 'isoformat'):
        return value.isoformat()
    return str(value).strip()


def read_rows(uploaded, expected_columns):
    """Read an .xlsx or .csv upload into a list of dicts keyed by header.

    Raises ImportError_ when the file cannot be read as a table at all, which
    is different from a row failing validation and is reported differently.
    """
    name = (uploaded.name or '').lower()

    if name.endswith('.csv'):
        text = uploaded.read().decode('utf-8-sig', errors='replace')
        table = list(csv.reader(io.StringIO(text)))
    elif name.endswith('.xlsx'):
        workbook = load_workbook(uploaded, read_only=True, data_only=True)
        table = [[_normalise(c) for c in row]
                 for row in workbook.active.iter_rows(values_only=True)]
    else:
        raise ImportError_('Upload a .xlsx or .csv file.')

    table = [row for row in table if any(_normalise(c) for c in row)]
    if not table:
        raise ImportError_('That file is empty.')

    header = [_normalise(c).lower().replace(' ', '_') for c in table[0]]
    missing = [c for c in expected_columns if c not in header]
    if missing:
        raise ImportError_(
            'Missing column(s): %s. Expected: %s.'
            % (', '.join(missing), ', '.join(expected_columns)))

    body = table[1:]
    if len(body) > MAX_ROWS:
        raise ImportError_('That file has %d rows; the limit is %d.'
                           % (len(body), MAX_ROWS))

    rows = []
    for values in body:
        padded = list(values) + [''] * (len(header) - len(values))
        rows.append({key: _normalise(value)
                     for key, value in zip(header, padded, strict=False)})
    return rows


def _to_form_data(row, mapping):
    return {field: row.get(column, '') for column, field in mapping.items()}


def validate(rows, kind):
    """Check every row, returning (forms, errors).

    `errors` is (row number, field, message), numbered as the spreadsheet shows
    them - header is row 1 - because that is what the person has open.
    """
    form_class = StudentForm if kind == 'students' else TeacherForm
    mapping = STUDENT_FIELDS if kind == 'students' else TEACHER_FIELDS
    key_field = 'USN' if kind == 'students' else 'id'

    forms_, errors = [], []
    seen = {}
    for index, row in enumerate(rows, start=2):
        form = form_class(_to_form_data(row, mapping))
        form.is_valid()

        # A file that repeats a key passes each row's uniqueness check on its
        # own - nothing is in the database yet - and then the second row
        # overwrites the first. Catch it here, where the whole file is visible.
        key = row.get('usn' if kind == 'students' else 'staff_id', '').strip().upper()
        if key and key in seen:
            errors.append((index, key_field,
                           'Duplicate of row %d in this file.' % seen[key]))
        elif key:
            seen[key] = index

        for field, messages in form.errors.items():
            for message in messages:
                errors.append((index, field, message))

        forms_.append(form)

    return forms_, errors


@transaction.atomic
def commit(forms_, kind):
    """Create every person in one transaction, returning their credentials.

    All or nothing: a half-imported intake with no record of which half is
    worse than a rejected file.
    """
    created = []
    for form in forms_:
        user, password = form.create_user()
        person = form.save(commit=False)
        person.user = user
        person.save()
        created.append({'name': person.name,
                        'key': person.USN if kind == 'students' else person.id,
                        'username': user.username,
                        'password': password})
    return created
